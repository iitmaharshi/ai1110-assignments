import openpyxl
from collections import Counter
import pandas as pd
from openpyxl.styles import Alignment

df = pd.read_excel('raw_data.xlsx', 'Sheet1', header=None)
a = df.to_numpy()
df = pd.DataFrame(a)

newlst = []
df.applymap(lambda x: newlst.append(x))
newlst.sort()

cou = Counter(newlst)
digits = list(cou.keys())
freq = list(cou.values())
lt = len(digits)
total = len(newlst)

df1 = pd.DataFrame()
df1['Digits'] = digits
df1['Frequency\n'] = freq
df1.to_excel('frequency.xlsx', index=False)
wb=openpyxl.load_workbook("frequency.xlsx")
ws=wb['Sheet1']
ws['A12']='Total'
ws['B12']=50
ws.column_dimensions['B'].width = 15
ws.row_dimensions[1].height = 30
ws['B1'].alignment = Alignment(wrap_text=True)

wb.save("frequency.xlsx")
wb.close()